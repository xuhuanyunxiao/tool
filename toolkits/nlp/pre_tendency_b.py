#!/usr/bin/python
# -*- coding:utf-8 -*-
from __future__ import division
import re
import jieba
from string import digits

stopwords = {}
stw = open("corpus/stop_words_tendency.txt", encoding='UTF-8')
for ws in stw:
    ws = ws.replace("\n", "")
    ws = ws.replace("\r", "")
    stopwords[ws] = 1
stw.close()

jieba.load_userdict('corpus/company.txt')
jieba.load_userdict('corpus/user_dict.txt')
jieba.load_userdict('corpus/bank_dict.txt')


def handle_contents(l_contents):
    lines = []
    for line in l_contents:
        lines.append(handle_content(line))
    return lines


def handle_content(content):
    content = str(content)
    raw = content.strip()
    line = ""
    if raw != "":
        word_list_1 = []
        line = ""
        remove_words = []
        raw = clear_sen(raw)
        word_list = filter(lambda x: len(x) > 0, map(etl, jieba.cut(raw, cut_all=False)))
        ll = list(word_list)
        for wd in ll:
            if wd in stopwords:
                remove_words.append(wd)

        for l in ll:
            if not (l in remove_words):
                word_list_1.append(l)

        for wd in word_list_1:
            line = line + wd + " "
    return line


def clear_sen(sent):
    sent = sent.replace("\n", "")
    sent = sent.replace('\r','')
    sent = sent.replace('\r\n','')
    reobj = re.compile('//@(.*?)[:\s]')
    sent = reobj.sub("", sent)
    reobj = re.compile("@(.*?)[:\s]")
    sent = reobj.sub("", sent)
    reobj = re.compile(r"\[[^\[\]]*?\]")
    sent = reobj.sub("", sent)

    sent = sent.replace("，", ",")
    sent = sent.replace("。", ".")
    sent = sent.replace("！", "!")
    sent = sent.replace("？", "?")
    reobj = re.compile("//(.*?)[:\s]")
    sent = reobj.sub("", sent)
    return sent


def etl(s):  # remove 标点和特殊字符
    regex = re.compile(r"[^\u4e00-\u9f5aa-zA-Z0-9]")
    s = regex.sub('', s)
    remove_digits = str.maketrans('', '', digits)
    res = s.translate(remove_digits)
    return res


def handle_test_excel(test_file, mid_file, all_file, tag_file, right_flag):
    # 从excel读取数据
    from openpyxl import Workbook
    from openpyxl import load_workbook
    test_corpus = []
    workbook = load_workbook(test_file)
    sheet_names = workbook.sheetnames  # 获得表单名字
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        for row in range(2, sheet.max_row + 1):
            row = row
            title_column = 1
            column_column = 2
            title = sheet.cell(row=row, column=title_column).value.encode('utf-8')
            content = sheet.cell(row=row, column=column_column).value.encode('utf-8')

            # python3 需要转换
            content = content.decode()

            words_str = handle_content(content)
            words_str = words_str.encode('utf-8')
            test_corpus.append((title, content, words_str))
    workbook.close()

    # 预测, 输出准确率
    import predict
    result = predict.predict_corpus([content_word[2] for content_word in test_corpus])
    print('测试文件: ', test_file)
    print('总条数: ', len(result))
    right_result = list(filter(lambda x: x == right_flag, result))
    print('正确条数; ', len(right_result))
    print('准确度: ', len(right_result) / len(result))

    # 将预处理的完毕的数据保存到中间文件中
    mid_wb = Workbook()
    mid_ws = mid_wb.active
    mid_ws.cell(row=1, column=1).value = 'words_str'
    for row in range(0, len(test_corpus)):
        mid_ws.cell(row=row+2, column=1).value = test_corpus[row][2]
    mid_wb.save(mid_file)
    print('保存中间数据: ', mid_file)

    # 保存所有数据
    all_wb = Workbook()
    all_ws = all_wb.active
    all_ws.cell(row=1, column=1).value = 'title'
    all_ws.cell(row=1, column=2).value = 'content'
    all_ws.cell(row=1, column=3).value = 'predict_flag'
    all_list = [content for content in zip(test_corpus, result)]
    for row in range(0, len(all_list)):
        all_ws.cell(row=row+2, column=1).value = all_list[row][0][0]
        all_ws.cell(row=row+2, column=2).value = all_list[row][0][1]
        all_ws.cell(row=row+2, column=3).value = all_list[row][1]
    all_wb.save(all_file)
    print('保存所有数据: ', all_file)

    # 将结果不正确的数据, 保存到tag_file中
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'title'
    ws.cell(row=1, column=2).value = 'content'
    error_list = [content[0] for content in zip(test_corpus, result) if content[1] != right_flag]
    for row in range(2, len(error_list)):
        ws.cell(row=row, column=1).value = error_list[row][0]
        ws.cell(row=row, column=2).value = error_list[row][1]
    wb.save(tag_file)
    print('保存错误数据: ', tag_file)


if __name__ == '__main__':
    handle_test_excel(u'test/0209保监会相关数据.xlsx', u'test/中间数据_0209保监会相关数据.xlsx', u'test/全部数据_0209保监会相关数据.xlsx', u'test/判断错误_0209保监会相关数据.xlsx', 1)
    # handle_test_excel(u'test/V2/0209保监会不相关数据.xlsx', u'test/V2/判断错误_0209保监会不相关数据.xlsx', 0)
